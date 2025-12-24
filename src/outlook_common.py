from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from playwright.sync_api import Locator, Page, TimeoutError as PlaywrightTimeoutError


OUTLOOK_MAIL_URL = "https://outlook.office.com/mail/0/?deeplink=mail%2F0%2F"


def _ts() -> str:
    return datetime.now().isoformat(timespec="seconds")


def log(message: str) -> None:
    print(f"[{_ts()}] {message}", flush=True)


def _mask_email(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    if "@" in value:
        local, domain = value.split("@", 1)
        local_mask = (local[:2] + "***") if len(local) >= 2 else "***"
        return f"{local_mask}@{domain}"
    return (value[:2] + "***") if len(value) >= 2 else "***"


def append_excel_summary(
    *,
    script_name: str,
    list_name: str,
    deleted_this_session: int,
    total_deleted: int,
    browser_name: str,
    headless: bool,
    email: str,
    excel_path: str | None = None,
) -> None:
    """Append one row to an Excel summary file.

    Designed for quick run-level totals across multiple scripts.
    """

    path = Path(excel_path) if excel_path else (Path(__file__).resolve().parent.parent / "outlook_delete_summary.xlsx")
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        log(f"Excel: openpyxl not available ({type(exc).__name__}: {exc})")
        return

    headers = [
        "timestamp",
        "script",
        "list",
        "deleted_this_session",
        "total_deleted",
        "browser",
        "headless",
        "account",
    ]

    row = [
        _ts(),
        script_name,
        list_name,
        int(deleted_this_session),
        int(total_deleted),
        browser_name,
        bool(headless),
        _mask_email(email),
    ]

    try:
        if path.exists():
            wb = load_workbook(path)
            ws = wb["Summary"] if "Summary" in wb.sheetnames else wb.active
            if ws.title != "Summary":
                ws.title = "Summary"
            # If file exists but is empty/no header, write header.
            if ws.max_row < 1:
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws.append(headers)

        ws.append(row)
        wb.save(path)
        log(f"Excel: appended summary row -> {path}")
    except PermissionError:
        log(f"Excel: cannot write (file open/locked): {path}")
    except Exception as exc:
        log(f"Excel: failed to write ({type(exc).__name__}: {exc})")


@dataclass(frozen=True)
class OutlookConfig:
    email: str
    password: str
    browser_name: str = "firefox"
    headless: bool = False
    timeout_ms: int = 30000


def load_config(*, browser_name: str | None = None, headless: bool | None = None, timeout_ms: int | None = None) -> OutlookConfig:
    email = (os.getenv("OUTLOOK_EMAIL") or "").strip()
    password = (os.getenv("OUTLOOK_PASSWORD") or "").strip()
    if not email or not password:
        raise RuntimeError("Missing OUTLOOK_EMAIL / OUTLOOK_PASSWORD. Set them in env or .env file.")

    env_browser = (os.getenv("PLAYWRIGHT_BROWSER") or "").strip().lower() or None
    env_headless = (os.getenv("PLAYWRIGHT_HEADLESS") or "").strip().lower() or None
    env_timeout = (os.getenv("PLAYWRIGHT_TIMEOUT_MS") or "").strip() or None

    resolved_browser = (browser_name or env_browser or "firefox").strip().lower()
    if resolved_browser not in {"chromium", "firefox", "webkit"}:
        raise ValueError("browser_name must be one of: chromium, firefox, webkit")

    resolved_headless: bool
    if headless is not None:
        resolved_headless = headless
    elif env_headless is not None:
        resolved_headless = env_headless in {"1", "true", "yes", "y"}
    else:
        resolved_headless = False

    resolved_timeout = timeout_ms
    if resolved_timeout is None and env_timeout is not None:
        try:
            resolved_timeout = int(env_timeout)
        except ValueError:
            resolved_timeout = None
    if resolved_timeout is None:
        resolved_timeout = 30000

    cfg = OutlookConfig(
        email=email,
        password=password,
        browser_name=resolved_browser,
        headless=resolved_headless,
        timeout_ms=resolved_timeout,
    )

    log(
        "Config loaded: "
        f"email={_mask_email(cfg.email)} browser={cfg.browser_name} "
        f"headless={cfg.headless} timeout_ms={cfg.timeout_ms}"
    )
    return cfg


def _robust_click(locator: Locator, *, timeout_ms: int, retries: int = 3) -> None:
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            locator.wait_for(state="visible", timeout=timeout_ms)
            # no_wait_after avoids hanging when the click triggers a slow/blocked navigation.
            locator.click(timeout=timeout_ms, force=(attempt >= retries - 1), no_wait_after=True)
            return
        except Exception as exc:
            last_err = exc

    if last_err:
        raise last_err


def _fill_first_visible(page: Page, selectors: list[str], value: str, *, timeout_ms: int) -> None:
    last_err: Exception | None = None
    for sel in selectors:
        loc = page.locator(sel).first
        try:
            loc.wait_for(state="visible", timeout=timeout_ms)
            loc.fill(value, timeout=timeout_ms)
            return
        except Exception as exc:
            last_err = exc
    if last_err:
        raise last_err


def _click_first_visible(page: Page, selectors: list[str], *, timeout_ms: int) -> None:
    last_err: Exception | None = None
    for sel in selectors:
        loc = page.locator(sel).first
        try:
            loc.wait_for(state="visible", timeout=timeout_ms)
            loc.click(timeout=timeout_ms, no_wait_after=True)
            return
        except Exception as exc:
            last_err = exc
    if last_err:
        raise last_err


def login(page: Page, email: str, password: str, *, timeout_ms: int) -> None:
    """Login to Microsoft/Outlook with common flow fallbacks."""
    log(f"Login: goto {OUTLOOK_MAIL_URL}")
    page.set_default_timeout(timeout_ms)
    page.goto(OUTLOOK_MAIL_URL, wait_until="domcontentloaded")

    # Sometimes Outlook shows a "Sign in" link first.
    try:
        sign_in_link = page.get_by_role("link", name="Sign in")
        if sign_in_link.is_visible(timeout=2000):
            log("Login: click 'Sign in' link")
            sign_in_link.click()
    except PlaywrightTimeoutError:
        pass

    # Handle "Pick an account" by selecting "Use another account".
    try:
        use_other = page.get_by_role("button", name="Use another account")
        if use_other.is_visible(timeout=1500):
            log("Login: 'Pick an account' -> click 'Use another account'")
            use_other.click()
    except PlaywrightTimeoutError:
        pass

    log(f"Login: fill email { _mask_email(email) }")

    _fill_first_visible(
        page,
        selectors=[
            "input[name='loginfmt']",
            "input[type='email']",
            "#i0116",
        ],
        value=email,
        timeout_ms=timeout_ms,
    )
    _click_first_visible(
        page,
        selectors=[
            "#idSIButton9",
            "button:has-text('Next')",
            "input[type='submit']",
        ],
        timeout_ms=timeout_ms,
    )

    log("Login: email submitted (Next)")

    # Ensure we actually advanced to password screen; if not, retry a bit.
    password_field = page.locator("input[name='passwd'], #i0118, input[type='password']").first
    try:
        password_field.wait_for(state="visible", timeout=min(timeout_ms, 8000))
    except PlaywrightTimeoutError:
        log("Login: password field not visible yet; retry Next/Enter")
        try:
            page.keyboard.press("Enter")
        except Exception:
            pass
        try:
            _click_first_visible(
                page,
                selectors=["#idSIButton9", "button:has-text('Next')", "input[type='submit']"],
                timeout_ms=min(timeout_ms, 8000),
            )
        except Exception:
            pass
        password_field.wait_for(state="visible", timeout=min(timeout_ms, 15000))

    log("Login: fill password")

    _fill_first_visible(
        page,
        selectors=[
            "input[name='passwd']",
            "input[type='password']",
            "#i0118",
        ],
        value=password,
        timeout_ms=timeout_ms,
    )
    _click_first_visible(
        page,
        selectors=[
            "#idSIButton9",
            "button:has-text('Sign in')",
            "input[type='submit']",
        ],
        timeout_ms=timeout_ms,
    )

    log("Login: password submitted (Sign in)")

    # Stay signed in prompt can appear with Yes/No or back button.
    try:
        log("Login: dismiss 'Stay signed in?' (Back)")
        page.locator("#idBtn_Back").click(timeout=4000)
    except PlaywrightTimeoutError:
        for btn_name in ("No", "Yes"):
            try:
                log(f"Login: dismiss 'Stay signed in?' ({btn_name})")
                page.get_by_role("button", name=btn_name).click(timeout=2000)
                break
            except PlaywrightTimeoutError:
                continue

    log(f"Login: done (url={page.url})")


def open_people(page: Page, *, timeout_ms: int) -> None:
    log("UI: open People")
    people_btn = page.locator("xpath=//button[@aria-label='People']").first
    people_btn.wait_for(state="visible", timeout=timeout_ms)
    try:
        pressed = people_btn.get_attribute("aria-pressed")
        if pressed and pressed.lower() == "true":
            log("UI: People already selected")
            return
    except Exception:
        pass

    _robust_click(people_btn, timeout_ms=timeout_ms, retries=3)


def open_contact_list(page: Page, list_name: str, *, timeout_ms: int) -> None:
    # Outlook UI varies; prefer left navigation items (treeitem/button/link) to avoid
    # matching non-clickable headers in the main pane.
    log(f"UI: open list '{list_name}'")

    selectors = [
        f"role=treeitem[name='{list_name}']",
        f"role=button[name='{list_name}']",
        f"role=link[name='{list_name}']",
        # Scoped to navigation first (more likely to be the left rail).
        f"xpath=//nav//*[self::span or self::div][normalize-space(.)='{list_name}']",
        f"xpath=//*[@role='navigation']//*[self::span or self::div][normalize-space(.)='{list_name}']",
        # Fallback: any visible text match.
        f"xpath=//span[normalize-space(.)='{list_name}']",
    ]

    _click_first_visible(page, selectors=selectors, timeout_ms=timeout_ms)


def click_delete_and_confirm(page: Page, *, timeout_ms: int, confirm_variant: str = "default") -> None:
    log("UI: click Delete")

    # Ensure something is selected; otherwise Outlook may show a dialog that can't proceed
    # or disable the Delete action.
    try:
        main = page.locator("div[role='main']").first
        first_checkbox = main.get_by_role("checkbox").first
        if first_checkbox.is_visible(timeout=1000):
            log("UI: select first checkbox")
            _robust_click(first_checkbox, timeout_ms=min(timeout_ms, 3000), retries=2)
    except Exception:
        pass

    delete_button = page.get_by_role("button", name="Delete").first
    delete_fallback = page.locator("xpath=//span[normalize-space(.)='Delete']").first

    try:
        _robust_click(delete_button, timeout_ms=timeout_ms, retries=3)
    except Exception:
        _robust_click(delete_fallback, timeout_ms=timeout_ms, retries=3)

    # Contact list deletion confirm button can be a plain <button> with direct text.
    # User-provided working selector:
    #   //div[contains(@class,'Dialog')]//button[text()='Delete']
    # Use normalize-space to handle whitespace.
    if confirm_variant == "contact_list":
        confirm_contact_list = page.locator(
            "xpath=//div[contains(@class,'Dialog')]//button[normalize-space(.)='Delete']"
        ).first
        try:
            log("UI: confirm Delete (contact_list Dialog button[text()='Delete'])")
            _robust_click(confirm_contact_list, timeout_ms=min(timeout_ms, 8000), retries=3)
            log("UI: delete confirmed")
            return
        except Exception:
            # Fall back to the default strategies.
            pass

    # Default confirm (Deleted Bin): keep ONLY this selector/behavior.
    confirm_span_exact = page.locator(
        "xpath=//div[contains(@class,'Dialog') or contains(@class,'ms-Dialog')]//button//span[normalize-space(.)='Delete']"
    ).first

    log("UI: confirm Delete (exact Dialog span xpath)")
    _robust_click(confirm_span_exact, timeout_ms=min(timeout_ms, 8000), retries=3)

    log("UI: delete confirmed")


def delete_flow(page: Page, *, list_name: str, timeout_ms: int, max_attempts: int = 10) -> None:
    """Open People, open a list, then delete+confirm with retries."""
    log(f"DeleteFlow: start list='{list_name}' max_attempts={max_attempts}")
    open_people(page, timeout_ms=timeout_ms)
    open_contact_list(page, list_name, timeout_ms=timeout_ms)

    last_err: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            log(f"DeleteFlow: attempt {attempt}/{max_attempts}")
            click_delete_and_confirm(page, timeout_ms=timeout_ms)
            log("DeleteFlow: success")
            return
        except Exception as exc:
            last_err = exc
            log(f"DeleteFlow: failed attempt {attempt} ({type(exc).__name__}: {exc})")
            try:
                log("DeleteFlow: reload")
                page.reload(wait_until="domcontentloaded")
            except Exception:
                pass
            try:
                open_people(page, timeout_ms=timeout_ms)
                open_contact_list(page, list_name, timeout_ms=timeout_ms)
            except Exception:
                pass

    raise RuntimeError(f"Delete flow failed after {max_attempts} attempts") from last_err


def delete_many(
    page: Page,
    *,
    list_name: str,
    timeout_ms: int,
    batch_size: int = 5,
    max_total: int | None = None,
    max_failures: int = 3,
    confirm_variant: str = "default",
    on_deleted: callable | None = None,
) -> int:
    """Delete repeatedly.

    Behavior: delete `batch_size` times, then reload and continue.
    Stops when:
    - `max_total` reached (if provided), OR
    - we fail `max_failures` times in a row (often means nothing left to delete).
    Returns the number of successful deletes.
    """
    if batch_size <= 0:
        raise ValueError("batch_size must be >= 1")

    log(f"DeleteMany: start list='{list_name}' batch_size={batch_size} max_total={max_total}")
    open_people(page, timeout_ms=timeout_ms)
    open_contact_list(page, list_name, timeout_ms=timeout_ms)

    total_deleted = 0
    consecutive_failures = 0

    while True:
        if max_total is not None and total_deleted >= max_total:
            log(f"DeleteMany: reached max_total={max_total}")
            return total_deleted

        try:
            click_delete_and_confirm(page, timeout_ms=timeout_ms, confirm_variant=confirm_variant)
            total_deleted += 1
            consecutive_failures = 0
            log(f"DeleteMany: deleted {total_deleted}")
            if on_deleted is not None:
                try:
                    on_deleted()
                except Exception:
                    # Never let progress hook break deletion.
                    pass
        except Exception as exc:
            consecutive_failures += 1
            log(f"DeleteMany: failure {consecutive_failures}/{max_failures} ({type(exc).__name__}: {exc})")
            # Try to recover UI state.
            try:
                page.reload(wait_until="domcontentloaded")
            except Exception:
                pass
            try:
                open_people(page, timeout_ms=timeout_ms)
                open_contact_list(page, list_name, timeout_ms=timeout_ms)
            except Exception:
                pass

            if consecutive_failures >= max_failures:
                # If we've never deleted anything, likely nothing selectable/left -> stop gracefully.
                if total_deleted == 0:
                    log("DeleteMany: stopping (too many failures; likely nothing left)")
                    return 0
                # Otherwise, signal the caller to restart the browser/login.
                raise RuntimeError("DeleteMany: too many consecutive failures; restart browser") from exc

        # After each batch, reload to refresh the list UI.
        if total_deleted > 0 and (total_deleted % batch_size == 0):
            log(f"DeleteMany: batch completed ({batch_size}); reload")
            try:
                page.reload(wait_until="domcontentloaded")
            except Exception:
                pass
            open_people(page, timeout_ms=timeout_ms)
            open_contact_list(page, list_name, timeout_ms=timeout_ms)
